import openpyxl
import sqlite3
import time
import re

DB_PATH = 'superpower.db'
XLSX_PATH = 'docs/添翼产品清单.xlsx'
VERSION_ID = 2

COL_MAP = {
    0: 'col_产品系统',
    1: 'col_应用角色',
    2: 'col_招标参数说明',
    3: 'col_功能说明',
    4: 'col_状态',
    5: 'col_业务分类',
    6: 'col_业务域',
    7: 'col_版本划分',
    8: 'col_远',
    9: 'col_交付工作量人月',
    10: 'col_控标点',
    11: 'col_控标点截图1',
    12: 'col_控标点截图2',
    13: 'col_控标点截图3',
    14: 'col_控标点文档说明',
    15: 'col_软著',
    16: 'col_备注',
    17: 'col_智慧医疗',
    18: 'col_智慧服务',
    19: 'col_智慧管理',
    20: 'col_互联互通',
    21: 'col_产品系统标识',
    22: 'col_模块标识',
    23: 'col_其他解决方案标记',
    24: 'col_文档维护人员',
    25: 'col_产品经理',
    26: 'col_父记录',
    27: 'col_内部版本',
    28: 'col_智能化',
    29: 'col_曜',
    30: 'col_驰',
    31: 'col_fy23',
    32: 'col_fy24',
    33: 'col_fy25',
    34: 'col_fy26',
    35: 'col_fy27',
    36: 'col_fy28',
    37: 'col_fy29',
    38: 'col_研发成本合计',
    39: 'col_销量曜',
    40: 'col_销量远',
    41: 'col_销量驰',
    42: 'col_出厂套价保本',
    43: 'col_负责人',
    44: 'col_产品线',
    45: 'col_资产类型',
}

def get_level(name):
    prefix = ''
    for ch in str(name):
        if ch.isdigit() or ch == '.':
            prefix += ch
        else:
            break
    return prefix.count('.') + 1 if prefix else 0

def parse_version_div(val):
    parts = []
    if val:
        s = str(val)
        if '曜' in s or 'A-曜' in s: parts.append('A-曜系列')
        if '远' in s or 'B-远' in s: parts.append('B-远系列')
        if '驰' in s or 'C-驰' in s: parts.append('C-驰系列')
        if not parts and s.strip():
            for seg in s.split():
                seg = seg.strip()
                if seg: parts.append(seg)
    return ' '.join(parts) if parts else None

def parse_bool(val):
    if val is None: return None
    s = str(val).strip()
    if s in ('是', 'true', '1', 'True'): return '是'
    if s in ('否', 'false', '0', 'False'): return '否'
    return s

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['筛选视图']
    
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    cur = conn.cursor()

    now_ms = int(time.time() * 1000)

    cur.execute("DELETE FROM data_entry WHERE version_id = ?", (VERSION_ID,))

    id_counter = [1000]

    def next_id():
        id_counter[0] += 1
        return id_counter[0]

    cat_map = {}
    domain_map = {}

    rows_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        level = get_level(name)
        if level < 3:
            continue
        biz_cat = str(row[5] or '').strip() if row[5] else ''
        biz_domain = str(row[6] or '').strip() if row[6] else ''

        fields = {}
        for col_idx, col_name in COL_MAP.items():
            if col_idx < len(row) and row[col_idx] is not None:
                val = row[col_idx]
                if col_idx == 7:
                    val = parse_version_div(val)
                elif col_idx in (8, 29, 30):
                    val = parse_bool(val)
                elif col_idx in (31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42):
                    try:
                        if isinstance(val, str):
                            val = float(val.replace(',', '')) if val.strip() else None
                        elif val is not None:
                            val = float(val)
                    except (ValueError, TypeError):
                        val = None
                else:
                    val = str(val).strip() if val else None
                if val is not None and val != '':
                    fields[col_name] = val

        fields['col_业务分类'] = biz_cat
        fields['col_业务域'] = biz_domain
        fields['col_产品系统'] = name

        rows_data.append({
            'name': name,
            'level': level,
            'biz_cat': biz_cat,
            'biz_domain': biz_domain,
            'fields': fields,
        })

    for rd in rows_data:
        cat = rd['biz_cat']
        if cat and cat not in cat_map:
            cat_id = next_id()
            cur.execute("""
                INSERT INTO data_entry (id, col_产品系统, col_业务分类, level, parent_id, version_id, sort_order, is_leaf, created_at, updated_at)
                VALUES (?, ?, ?, 1, NULL, ?, ?, 0, ?, ?)
            """, (cat_id, cat, cat, VERSION_ID, len(cat_map), now_ms, now_ms))
            cat_map[cat] = cat_id

    for rd in rows_data:
        domain = rd['biz_domain']
        cat = rd['biz_cat']
        if domain and domain not in domain_map:
            parent_id = cat_map.get(cat)
            domain_id = next_id()
            cur.execute("""
                INSERT INTO data_entry (id, col_产品系统, col_业务分类, col_业务域, level, parent_id, version_id, sort_order, is_leaf, created_at, updated_at)
                VALUES (?, ?, ?, ?, 2, ?, ?, ?, 0, ?, ?)
            """, (domain_id, domain, cat, domain, parent_id, VERSION_ID, len(domain_map), now_ms, now_ms))
            domain_map[domain] = domain_id

    parent_stack = {}
    level3_sort = {}

    for rd in rows_data:
        level = rd['level']
        name = rd['name']
        fields = rd['fields']
        entry_id = next_id()

        if level == 3:
            domain = rd['biz_domain']
            parent_id = domain_map.get(domain)
            sort_key = domain
            if sort_key not in level3_sort:
                level3_sort[sort_key] = 0
            sort_order = level3_sort[sort_key]
            level3_sort[sort_key] += 1
            parent_stack[3] = entry_id
        else:
            parent_id = parent_stack.get(level - 1)
            parent_stack[level] = entry_id
            sort_order = 0
            if parent_id:
                key = f"{parent_id}_{level}"
                if key not in level3_sort:
                    level3_sort[key] = 0
                sort_order = level3_sort[key]
                level3_sort[key] += 1

        cols = ['id', 'level', 'parent_id', 'version_id', 'sort_order', 'is_leaf', 'created_at', 'updated_at']
        vals = [entry_id, level, parent_id, VERSION_ID, sort_order, 1, now_ms, now_ms]

        for col_name, val in fields.items():
            cols.append(col_name)
            vals.append(val)

        placeholders = ','.join(['?'] * len(cols))
        col_str = ','.join(cols)
        cur.execute(f"INSERT INTO data_entry ({col_str}) VALUES ({placeholders})", vals)

        if parent_id:
            cur.execute("UPDATE data_entry SET is_leaf = 0 WHERE id = ? AND version_id = ?", (parent_id, VERSION_ID))

    conn.commit()

    cur.execute("SELECT level, COUNT(*) FROM data_entry WHERE version_id = ? GROUP BY level ORDER BY level", (VERSION_ID,))
    for row in cur.fetchall():
        print(f"  Level {row[0]}: {row[1]}")

    cur.execute("SELECT COUNT(*) FROM data_entry WHERE version_id = ?", (VERSION_ID,))
    print(f"  Total: {cur.fetchone()[0]}")

    cur.execute("""
        SELECT COUNT(*) FROM data_entry c
        JOIN data_entry p ON c.parent_id = p.id
        WHERE c.version_id = ? AND p.version_id = ? AND c.level != p.level + 1
    """, (VERSION_ID, VERSION_ID))
    print(f"  Level errors: {cur.fetchone()[0]}")

    cur.execute("""
        SELECT COUNT(*) FROM data_entry
        WHERE version_id = ? AND parent_id IS NOT NULL
        AND parent_id NOT IN (SELECT id FROM data_entry WHERE version_id = ?)
    """, (VERSION_ID, VERSION_ID))
    print(f"  Orphan refs: {cur.fetchone()[0]}")

    cur.execute("""
        SELECT COUNT(*) FROM (
            SELECT parent_id, TRIM(col_产品系统) as name FROM data_entry
            WHERE version_id = ? AND level >= 3
            GROUP BY parent_id, TRIM(col_产品系统) HAVING COUNT(*) > 1
        )
    """, (VERSION_ID,))
    print(f"  Duplicates: {cur.fetchone()[0]}")

    conn.close()
    print("Done!")

if __name__ == '__main__':
    main()
