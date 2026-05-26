import openpyxl
import requests
import re

BASE = "http://localhost:8080/api"
XLSX = "/Users/craneking/workspace/工程设计/superPowerTest/docs/添翼产品清单.xlsx"

# 登录
resp = requests.post(f"{BASE}/auth/login", json={"username": "admin", "password": "123456"})
token = resp.json()["data"]["token"]
headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

# 获取编辑中的版本
resp = requests.get(f"{BASE}/versions", headers=headers)
draft = [v for v in resp.json()["data"] if v["status"] == "draft"]
version_id = draft[0]["id"]
print(f"使用版本 ID: {version_id}")

# 获取该分类下所有已导入的条目，建立 产品名 → ID 映射
resp = requests.get(f"{BASE}/data/query/{version_id}", headers=headers)
all_entries = resp.json()["data"]
existing = {}
for e in all_entries:
    if e.get("colProductSystem"):
        name = e["colProductSystem"].strip()
        existing[name] = e["id"]

print(f"已有 {len(existing)} 条已导入条目")

# 字段映射（Excel列号 → API字段名）
field_map = {
    1: "colProductSystem", 2: "colAppRole", 3: "colBidParamDesc",
    4: "colFeatureDesc", 5: "colStatus", 6: "colBizCategory",
    7: "colBizDomain", 8: "colVersionDivision", 9: "colYuan",
    10: "colDeliveryWorkload", 11: "colControlPoint",
    12: "colControlPointImg1", 13: "colControlPointImg2",
    14: "colControlPointImg3", 15: "colControlPointDoc",
    16: "colCopyright", 17: "colRemark",
    18: "colSmartMedical", 19: "colSmartService",
    20: "colSmartManagement", 21: "colInterconnection",
    22: "colProductSysId", 23: "colModuleId",
    24: "colOtherSolutionTag", 25: "colDocMaintainer",
    26: "colProductManager", 27: "colParentRecord",
    28: "colInternalVersion", 29: "colIntelligent",
    30: "colYao", 31: "colChi",
    32: "colFY23", 33: "colFY24", 34: "colFY25",
    35: "colFY26", 36: "colFY27", 37: "colFY28",
    38: "colFY29", 39: "colRDCostTotal",
    40: "colSalesYao", 41: "colSalesYuan", 42: "colSalesChi",
    43: "colFactoryPrice", 44: "colPrincipal",
    45: "colProductLine", 46: "colAssetType"
}

def get_prefix(product_name):
    """提取产品名中的编号前缀，如 '1.1.1.1'"""
    if not product_name:
        return None
    m = re.match(r'^([\d.]+)', product_name.strip())
    return m.group(1).rstrip('.') if m else None

def get_parent_prefix(prefix):
    """获取父级编号前缀（去掉最后一段）"""
    parts = prefix.split('.')
    if len(parts) <= 3:
        return None
    return '.'.join(parts[:-1])

def find_parent_id(product_name, existing):
    """通过产品名中的编号前缀查找父级ID"""
    prefix = get_prefix(product_name)
    if not prefix:
        return None
    parent_prefix = get_parent_prefix(prefix)
    if not parent_prefix:
        return None
    # 查找匹配的父级产品名称
    for name, eid in existing.items():
        if name.startswith(parent_prefix + ' '):
            return eid
    return None

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

imported = 0
skipped = 0
no_parent = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row=row, column=6).value
    if cat != '1. 数智底座-数据':
        continue

    product = ws.cell(row=row, column=1).value
    if not product:
        continue

    product_name = product.strip()

    # 跳过已导入的（包括已有的L3）
    if product_name in existing:
        skipped += 1
        continue

    # 查找父级
    parent_id = find_parent_id(product_name, existing)

    # 确定层级
    prefix = get_prefix(product_name)
    level = prefix.count('.') + 1 if prefix else 4

    payload = {
        "versionId": version_id,
        "parentId": parent_id,
        "level": level,
        "sortOrder": 0,
    }

    for col_idx, field in field_map.items():
        val = ws.cell(row=row, column=col_idx).value
        if val is not None:
            payload[field] = val

    resp = requests.post(f"{BASE}/data", json=payload, headers=headers)
    if resp.status_code == 200:
        entry = resp.json()["data"]
        existing[product_name] = entry["id"]
        imported += 1
        if imported <= 5:
            print(f'  ✅ L{level} {product_name} (parent={parent_id})')
    else:
        print(f'  ❌ {product_name}: {resp.text}')

print(f"\n✅ 导入完成: 成功 {imported} 条, 已有 {skipped} 条跳过")
