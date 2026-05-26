import openpyxl
import requests
import json

BASE = "http://localhost:8080/api"
XLSX = "/Users/craneking/workspace/工程设计/superPowerTest/docs/添翼产品清单.xlsx"

# 登录
resp = requests.post(f"{BASE}/auth/login", json={"username": "admin", "password": "123456"})
token = resp.json()["data"]["token"]
headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

# 获取版本
resp = requests.get(f"{BASE}/versions", headers=headers)
versions = resp.json()["data"]
print(f"可用版本: {[v['versionNo'] for v in versions]}")

# 使用编辑中的版本
draft = [v for v in versions if v["status"] == "draft"]
version_id = draft[0]["id"] if draft else versions[-1]["id"]
print(f"使用版本 ID: {version_id} (status: {'draft' if draft else 'released'})")

# 读取 Excel
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# 收集 L1 和 L2 数据，保留顺序
seen_cats = []
cat_domains = {}

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row=row, column=6).value
    domain = ws.cell(row=row, column=7).value
    if cat and cat not in seen_cats:
        seen_cats.append(cat)
    if cat and domain:
        if cat not in cat_domains:
            cat_domains[cat] = []
        if domain not in cat_domains[cat]:
            cat_domains[cat].append(domain)

# 创建 L1（业务分类）
print("\n=== 创建业务分类 (L1) ===")
cat_id_map = {}
for i, cat in enumerate(seen_cats):
    payload = {
        "versionId": version_id,
        "parentId": None,
        "level": 1,
        "sortOrder": i,
        "colBizCategory": cat,
        "colProductSystem": cat
    }
    resp = requests.post(f"{BASE}/data", json=payload, headers=headers)
    if resp.status_code == 200:
        entry = resp.json()["data"]
        cat_id_map[cat] = entry["id"]
        print(f"  ✅ {cat} → ID {entry['id']}")
    else:
        print(f"  ❌ {cat}: {resp.text}")

# 创建 L2（业务域）
print("\n=== 创建业务域 (L2) ===")
for cat, domains in cat_domains.items():
    parent_id = cat_id_map.get(cat)
    if not parent_id:
        continue
    for i, domain in enumerate(domains):
        payload = {
            "versionId": version_id,
            "parentId": parent_id,
            "level": 2,
            "sortOrder": i,
            "colBizDomain": domain,
            "colProductSystem": domain
        }
        resp = requests.post(f"{BASE}/data", json=payload, headers=headers)
        if resp.status_code == 200:
            entry = resp.json()["data"]
            print(f"  ✅ {cat} → {domain} → ID {entry['id']}")
        else:
            print(f"  ❌ {cat} → {domain}: {resp.text}")

print("\n✅ 导入完成")
