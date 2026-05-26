import openpyxl
import requests

BASE = "http://localhost:8080/api"
XLSX = "/Users/craneking/workspace/工程设计/superPowerTest/docs/添翼产品清单.xlsx"

# 登录
resp = requests.post(f"{BASE}/auth/login", json={"username": "admin", "password": "123456"})
token = resp.json()["data"]["token"]
headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

# 获取编辑中的版本
resp = requests.get(f"{BASE}/versions", headers=headers)
draft = [v for v in resp.json()["data"] if v["status"] == "draft"]
version_id = draft[0]["id"]
print(f"使用版本 ID: {version_id}")

# 获取 L2 节点，建立 (业务分类, 业务域) → ID 映射
resp = requests.get(f"{BASE}/data/tree/{version_id}", headers=headers)
l2_map = {}
for l1 in resp.json()["data"]:
    cat_label = l1["label"]
    for l2 in l1.get("children", []):
        key = (cat_label, l2["label"])
        l2_map[key] = l2["id"]

print(f"已加载 {len(l2_map)} 个 L2 节点")

# 读取 Excel
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

field_map = {
    1: "colProductSystem", 2: "colAppRole", 3: "colBidParamDesc",
    4: "colFeatureDesc", 5: "colStatus", 6: "colBizCategory",
    7: "colBizDomain", 8: "colVersionDivision", 9: "colYuan",
    10: "colDeliveryWorkload", 11: "colControlPoint",
    12: "colControlPointImg1", 13: "colControlPointImg2",
    14: "colControlPointImg3", 15: "colControlPointDoc",
    16: "colCopyright", 17: "colRemark",
    18: "colSmartMedical", 19: "colSmartService",
    20: "colSmartManagement", 21: "colInterconnection",
    22: "colProductSysId", 23: "colModuleId",
    24: "colOtherSolutionTag", 25: "colDocMaintainer",
    26: "colProductManager", 27: "colParentRecord",
    28: "colInternalVersion", 29: "colIntelligent",
    30: "colYao", 31: "colChi",
    32: "colFY23", 33: "colFY24", 34: "colFY25",
    35: "colFY26", 36: "colFY27", 37: "colFY28",
    38: "colFY29", 39: "colRDCostTotal",
    40: "colSalesYao", 41: "colSalesYuan", 42: "colSalesChi",
    43: "colFactoryPrice", 44: "colPrincipal",
    45: "colProductLine", 46: "colAssetType"
}

imported = 0
skipped = 0

for row in range(2, ws.max_row + 1):
    flag = ws.cell(row=row, column=22).value
    if str(flag) != "1":
        continue

    cat = ws.cell(row=row, column=6).value or ""
    domain = ws.cell(row=row, column=7).value or ""
    parent_id = l2_map.get((cat, domain))

    if not parent_id:
        print(f"  ⚠ 未找到L2: ({cat}, {domain})")
        skipped += 1
        continue

    payload = {
        "versionId": version_id,
        "parentId": parent_id,
        "level": 3,
        "sortOrder": 0
    }

    for col_idx, field in field_map.items():
        val = ws.cell(row=row, column=col_idx).value
        if val is not None:
            payload[field] = val

    resp = requests.post(f"{BASE}/data", json=payload, headers=headers)
    if resp.status_code == 200:
        imported += 1
    else:
        print(f"  ❌ 导入失败 row {row}: {resp.text}")

print(f"\n✅ 导入完成: 成功 {imported} 条, 跳过 {skipped} 条")
